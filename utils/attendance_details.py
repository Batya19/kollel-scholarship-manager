import pandas as pd
from datetime import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.settings import NINE_AM_MINUTES
from utils.time_utils import time_to_minutes, calculate_hours_between, get_session_config
from utils.date_utils import get_hebrew_day_name


def _format_time_field(
        time_value,
        ws,
        row: int,
        column: int,
        allow_zero: bool = False):
    """
    Format and set a time field in worksheet.

    Args:
        time_value: Time value to format
        ws: Worksheet object
        row (int): Row number
        column (int): Column number
        allow_zero (bool): Whether time(0,0) is valid (False means it's missing)
    """
    if pd.isna(time_value) or time_value is pd.NaT:
        ws.cell(row=row, column=column).value = "חסר"
        ws.cell(row=row, column=column).font = Font(color="FF0000")
        return None
    elif isinstance(time_value, time):
        if not allow_zero and time_value == time(0, 0):
            ws.cell(row=row, column=column).value = "חסר"
            ws.cell(row=row, column=column).font = Font(color="FF0000")
            return None
        ws.cell(row=row, column=column).value = time_value.strftime('%H:%M')
        return time_value
    elif hasattr(time_value, 'time'):
        ws.cell(row=row, column=column).value = time_value.time().strftime('%H:%M')
        return time_value.time()
    else:
        ws.cell(row=row, column=column).value = str(time_value)
        return None


def _calculate_daily_bonuses(student_data: pd.DataFrame) -> dict:
    """
    Calculate daily bonuses for each date and session type.

    Args:
        student_data (pd.DataFrame): Student attendance data

    Returns:
        dict: Dictionary mapping (date, session_type) to bonus string
    """
    from collections import defaultdict
    daily_bonuses = defaultdict(lambda: "-")

    student_data_grouped = student_data.copy()
    student_data_grouped = student_data_grouped.sort_values(['תאריך', 'סדר'])

    for session_type in ['בוקר', 'צהריים']:
        session_data = student_data_grouped[student_data_grouped['סדר']
                                            == session_type]
        if session_data.empty:
            continue

        config = get_session_config(session_type)

        daily_grouped = session_data.groupby('תאריך').agg({
            'שעת כניסה': 'min',
            'שעת יציאה': 'max',
            'רצופות': lambda x: 'כן' in x.values if not x.empty else False
        })

        for date_val, row_data in daily_grouped.iterrows():
            entry_time = row_data['שעת כניסה']
            exit_time = row_data['שעת יציאה']
            is_continuous = row_data['רצופות']

            if isinstance(
                entry_time,
                time) and isinstance(
                exit_time,
                time) and exit_time != time(
                0,
                    0):
                entry_minutes = time_to_minutes(entry_time)
                exit_minutes = time_to_minutes(exit_time)
                perfect_start_minutes = time_to_minutes(
                    config['PERFECT_START'])
                end_minutes = time_to_minutes(config['END'])

                is_perfect = (
                    entry_minutes <= perfect_start_minutes and
                    exit_minutes >= end_minutes and
                    is_continuous
                )

                if is_perfect:
                    daily_bonuses[(date_val, session_type)] = "20 ₪"
                elif is_continuous:
                    daily_bonuses[(date_val, session_type)] = "5 ₪"

    return daily_bonuses


def _add_sheet_header(ws, student_name: str, student_id: str):
    """
    Add header to student detail sheet.

    Args:
        ws: Worksheet object
        student_name (str): Student full name
        student_id (str): Student ID number
    """
    ws.merge_cells('A1:M1')
    ws['A1'] = f"פירוט נוכחות - {student_name}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = f"מספר זהות: {student_id}"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal='center')


def _add_column_headers(ws, row: int):
    """
    Add column headers to worksheet.

    Args:
        ws: Worksheet object
        row (int): Row number for headers
    """
    headers = [
        'תאריך', 'יום', 'סדר', 'שעת כניסה', 'שעת יציאה',
        'רצופות', 'סך שעות', 'סטטוס', 'זמן איחור (דק\')',
        'הגיע לפני 9:00?', 'בונוס יומי', 'שעות חסרות', 'הערות'
    ]

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(
            start_color="D3D3D3",
            end_color="D3D3D3",
            fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def _add_summary_section(ws, summary_row: int, result: dict):
    """
    Add summary section to worksheet.

    Args:
        ws: Worksheet object
        summary_row (int): Starting row for summary
        result (dict): Scholarship calculation results
    """
    ws.merge_cells(f'A{summary_row}:M{summary_row}')
    ws[f'A{summary_row}'] = "סיכום"
    ws[f'A{summary_row}'].font = Font(size=14, bold=True)
    ws[f'A{summary_row}'].alignment = Alignment(horizontal='center')
    ws[f'A{summary_row}'].fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{summary_row}'].font = Font(size=14, bold=True, color="FFFFFF")

    summary_row += 2

    summary_data = [
        ("ימי נוכחות - בוקר:", result.get('בוקר_attended_days', 0)),
        ("ימי נוכחות - צהריים:", result.get('צהריים_attended_days', 0)),
        ("סה\"כ איחורים - בוקר:", result.get('בוקר_late_days', 0)),
        ("סה\"כ איחורים - צהריים:", result.get('צהריים_late_days', 0)),
        ("סה\"כ היעדרויות - בוקר:", result.get('בוקר_absent_days', 0)),
        ("סה\"כ היעדרויות - צהריים:", result.get('צהריים_absent_days', 0)),
        ("סה\"כ שעות - בוקר:", result.get('בוקר_total_hours', 0)),
        ("סה\"כ שעות - צהריים:", result.get('צהריים_total_hours', 0)),
        ("", ""),
        ("מלגת בסיס:", f"{result.get('מלגת בסיס', 0)} ₪"),
        ("", ""),
        ("תוספות:", ""),
        ("  בונוס יומי - בוקר:", f"{result.get('בונוס יומי בוקר', 0)} ₪"),
        ("  בונוס יומי - צהריים:", f"{result.get('בונוס יומי צהריים', 0)} ₪"),
        ("סה\"כ תוספות:", f"{result.get('תוספות', 0)} ₪"),
        ("", ""),
        ("בונוסים חודשיים:", ""),
        ("  תוספת דרגה 1:", f"{result.get('תוספת דרגה 1', 0)} ₪"),
        ("  תוספת דרגה 2:", f"{result.get('תוספת דרגה 2', 0)} ₪"),
        ("  נוכחות מושלמת:", f"{result.get('נוכחות מושלמת', 0)} ₪"),
        ("  הגעה מוקדמת:", f"{result.get('הגעה מוקדמת', 0)} ₪"),
        ("", ""),
        ("סך הכל:", f"{result.get('סך הכל', 0)} ₪"),
    ]

    for label, value in summary_data:
        cell = ws.cell(row=summary_row, column=1)
        cell.value = label

        if label == "תוספות:":
            cell.font = Font(bold=True, size=12)
        elif label.startswith("  "):
            cell.font = Font(bold=False)
        else:
            cell.font = Font(bold=True)

        ws.cell(row=summary_row, column=2).value = value
        summary_row += 1


def _set_column_widths(ws):
    """
    Set column widths for worksheet.

    Args:
        ws: Worksheet object
    """
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 20


def add_detailed_sheets(
        excel_file: str,
        input_data: pd.DataFrame,
        scholarship_results: list,
        working_days: int):
    """
    מוסיף גליונות מפורטים לכל אברך בקובץ Excel קיים.

    Args:
        excel_file (str): נתיב לקובץ Excel (שכבר מכיל את גליון הסיכום)
        input_data (pd.DataFrame): נתוני הנוכחות המקוריים
        scholarship_results (list): תוצאות חישוב המלגות
        working_days (int): מספר ימי עבודה בחודש
    """
    
    # Load existing file
    wb = load_workbook(excel_file)
    
    # Rename first sheet to "Summary"
    if 'Sheet' in wb.sheetnames or wb.sheetnames[0]:
        ws_summary = wb.active
        ws_summary.title = "סיכום"
        # Set right-to-left direction (RTL)
        ws_summary.sheet_view.rightToLeft = True
    
    # Loop through each student and create a sheet
    for result in scholarship_results:
        student_id = result['מספר זהות']
        student_name = result['שם מלא']
        
        # Filter student data
        student_data = input_data[input_data['זהות'] == student_id].copy()
        
        if student_data.empty:
            continue
        
        # Create new sheet for student (limited to 31 chars - Excel limitation)
        sheet_name = student_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Set right-to-left direction (RTL)
        ws.sheet_view.rightToLeft = True
        
        # Add header
        _add_sheet_header(ws, student_name, student_id)
        
        # Add column headers (row 4)
        _add_column_headers(ws, 4)
        
        # Add attendance data
        # Calculate daily bonuses grouped by date + session (as in actual calculation)
        daily_bonuses = _calculate_daily_bonuses(student_data)
        
        # Track date+session already displayed (to show bonus only once)
        displayed_dates = set()

        row = 5
        for _, record in student_data.iterrows():
            # Date
            ws.cell(row=row, column=1).value = record['תאריך'].strftime(
                '%d/%m/%Y') if hasattr(record['תאריך'], 'strftime') else str(record['תאריך'])

            # Day of week
            if hasattr(record['תאריך'], 'weekday'):
                ws.cell(
                    row=row, column=2).value = get_hebrew_day_name(
                    record['תאריך'].weekday())

            # Session type
            session_type = record['סדר']
            ws.cell(row=row, column=3).value = session_type

            # Entry time
            entry_time = _format_time_field(
                record['שעת כניסה'], ws, row, 4, allow_zero=True)

            # Exit time
            exit_time = _format_time_field(
                record['שעת יציאה'], ws, row, 5, allow_zero=False)

            # Continuous
            ws.cell(row=row, column=6).value = record.get('רצופות', '')

            # Total hours - calculated
            config = get_session_config(session_type)

            if isinstance(
                entry_time,
                time) and isinstance(
                exit_time,
                time) and exit_time != time(
                0,
                    0):
                hours = calculate_hours_between(
                    entry_time, exit_time, config['START'], config['END'])
                ws.cell(row=row, column=7).value = hours
            else:
                ws.cell(row=row, column=7).value = 0

            # Status (late/on-time)
            if isinstance(entry_time, time):
                entry_minutes = time_to_minutes(entry_time)
                late_threshold_minutes = time_to_minutes(
                    config['LATE_THRESHOLD'])
                very_late_threshold_minutes = time_to_minutes(
                    config['VERY_LATE_THRESHOLD'])

                if entry_minutes >= very_late_threshold_minutes:
                    ws.cell(row=row, column=8).value = "❌ איחור משמעותי"
                    ws.cell(row=row, column=8).font = Font(color="FF0000")
                    ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
                elif entry_minutes >= late_threshold_minutes:
                    ws.cell(row=row, column=8).value = "⚠️ איחור"
                    ws.cell(row=row, column=8).font = Font(color="FFA500")
                    ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
                else:
                    ws.cell(row=row, column=8).value = "✅ בזמן"
                    ws.cell(row=row, column=8).font = Font(color="008000")
                    ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')

            # Late minutes
            if isinstance(entry_time, time):
                entry_minutes_calc = time_to_minutes(entry_time)
                perfect_start_minutes = time_to_minutes(
                    config['PERFECT_START'])

                if entry_minutes_calc > perfect_start_minutes:
                    late_minutes = entry_minutes_calc - perfect_start_minutes
                    ws.cell(row=row, column=9).value = late_minutes
                else:
                    ws.cell(row=row, column=9).value = 0

            # Arrived before 9:00? (morning only)
            if session_type == 'בוקר' and isinstance(entry_time, time):
                entry_minutes_calc = time_to_minutes(entry_time)
                if entry_minutes_calc < NINE_AM_MINUTES:
                    ws.cell(row=row, column=10).value = "כן"
                    ws.cell(row=row, column=10).font = Font(color="008000")
                else:
                    ws.cell(row=row, column=10).value = "לא"
            else:
                ws.cell(row=row, column=10).value = "-"

            # Daily bonus - display only in first row of each date+session
            date_key = (record['תאריך'], session_type)
            if date_key not in displayed_dates:
                ws.cell(
                    row=row,
                    column=11).value = daily_bonuses.get(
                    date_key,
                    "-")
                displayed_dates.add(date_key)
            else:
                # Additional rows for same day - bonus already counted above
                ws.cell(row=row, column=11).value = "↑ כלול"
                ws.cell(
                    row=row,
                    column=11).font = Font(
                    color="808080",
                    italic=True)

            # Missing hours
            expected_hours = config['HOURS']
            actual_hours = ws.cell(row=row, column=7).value or 0
            missed_hours = round(expected_hours - actual_hours, 2)
            ws.cell(
                row=row,
                column=12).value = missed_hours if missed_hours > 0 else 0

            # Notes
            notes = []
            if exit_time == time(0, 0):
                notes.append("חסר זמן יציאה")
            if isinstance(entry_time, time):
                entry_minutes_calc = time_to_minutes(entry_time)
                very_late_threshold_minutes = time_to_minutes(
                    config['VERY_LATE_THRESHOLD'])
                if entry_minutes_calc >= very_late_threshold_minutes:
                    notes.append("איחור חמור")

            ws.cell(row=row, column=13).value = ", ".join(
                notes) if notes else ""

            row += 1

        # Add summary at the bottom
        summary_row = row + 2
        _add_summary_section(ws, summary_row, result)

        # Adjust column widths
        _set_column_widths(ws)

    # Save the file
    wb.save(excel_file)
