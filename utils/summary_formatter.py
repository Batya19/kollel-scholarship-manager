from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def format_summary_sheet(excel_file: str):
    """
    Format the summary sheet in a beautiful and readable way.
    
    Args:
        excel_file (str): Path to Excel file
    """
    wb = load_workbook(excel_file)
    ws = wb['סיכום'] if 'סיכום' in wb.sheetnames else wb.active

    # Find header row (row 1)
    header_row = 1

    # Format header row - classic and simple style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format headers - simple and classic
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)

        # Minimalist styling
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Light gray
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Header row height - higher for wrapped text
    ws.row_dimensions[header_row].height = 40

    # Format data cells - simple and classic
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            # Center alignment for numbers
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # Adjust column widths
    column_widths = {
        'מספר זהות': 14,
        'שם מלא': 25,
        'התראות': 50,
        'מלגת בסיס': 14,
        'תוספות': 12,
        'בונוס יומי בוקר': 16,
        'בונוס יומי צהריים': 18,
        'סך הכל': 12,
        'ציון חבורה': 13,
        'חבורה': 12,
        'ציון מוסר': 12,
        'מוסר': 12,
        'ציון סיכומים': 14,
        'סיכומים': 12,
        'ציון מבחן הלכה': 16,
        'מבחן הלכה': 14,
        'ציון מבחן שס': 15,
        'מבחן שס': 13,
        'תוספת דרגה 1': 16,
        'אנס דרגה 1': 18,
        'תוספת דרגה 2': 16,
        'אנס דרגה 2': 18,
        'נוכחות מושלמת': 16,
        'אנס נוכחות מושלמת': 20,
        'הגעה מוקדמת': 15,
        'אנס הגעה מוקדמת': 20,
        'סך סופי': 12,
    }
    
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        col_letter = get_column_letter(col)
        
        if header in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[header]
        else:
            # Default width - wider so text is readable
            # Calculate minimum width based on header length
            if header:
                # Calculate minimum width: header length + 2 chars padding
                min_width = max(15, len(str(header)) + 2)
                ws.column_dimensions[col_letter].width = min_width
            else:
                ws.column_dimensions[col_letter].width = 15

    # Freeze header row
    ws.freeze_panes = ws[f'A{header_row + 1}']

    # Save
    wb.save(excel_file)
