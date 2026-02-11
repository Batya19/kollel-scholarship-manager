from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from models.scholarship import KollelScholarship
from utils.attendance_details import add_detailed_sheets
from utils.summary_formatter import format_summary_sheet
from utils.date_utils import get_hebrew_month_name


def _build_grade_formula(grade_col: str, row: int, max_value: int) -> str:
    """
    Build formula for grade calculation.

    Args:
        grade_col (str): Grade column letter
        row (int): Row number
        max_value (int): Maximum value for this grade

    Returns:
        str: Excel formula
    """
    formula = (
        f'IF(ISNUMBER({grade_col}{row}),'
        f'IF({grade_col}{row}>=96,{max_value},'
        f'ROUND({max_value}*{grade_col}{row}/100,0)),"")'
    )
    return formula


def _build_bonus_with_override_formula(
        base_col: str,
        override_col: str,
        row: int,
        default_value: int) -> str:
    """
    Build formula for bonus with optional manual override.
    
    This calculates the DIFFERENCE to add/subtract from the total,
    not the full bonus amount (which is already included in "סך הכל").

    Args:
        base_col (str): Base bonus column letter
        override_col (str): Override reason column letter
        row (int): Row number
        default_value (int): Default bonus value when overridden

    Returns:
        str: Excel formula that returns the difference to apply
    """
    # If override is set: return (new_value - original_value)
    # If no override: return 0 (no change needed)
    return (
        f'IF(AND({override_col}{row}<>"",{override_col}{row}<>0),'
        f'IF(ISNUMBER({base_col}{row}),IF({base_col}{row}=0,{default_value},{base_col}{row}),{default_value})'
        f'-IF(ISNUMBER({base_col}{row}),{base_col}{row},0),'
        f'0)')


def _build_final_total_formula(
        total_col: str,
        grade_cols: dict,
        bonus_cols: dict,
        row: int) -> str:
    """
    Build final total formula combining all components.

    Args:
        total_col (str): Base total column letter
        grade_cols (dict): Dictionary of grade column letters
        bonus_cols (dict): Dictionary of bonus and override column letters
        row (int): Row number

    Returns:
        str: Complete Excel formula
    """
    # Base
    parts = [f'{total_col}{row}']

    # Grades
    for col in grade_cols.values():
        parts.append(f'IF(ISNUMBER({col}{row}),{col}{row},0)')

    # Bonuses with manual override option
    bonus_overrides = [
        (bonus_cols['tier1'], bonus_cols['tier1_reason'], 190),
        (bonus_cols['tier2'], bonus_cols['tier2_reason'], 200),
        (bonus_cols['perfect'], bonus_cols['perfect_reason'], 200),
        (bonus_cols['early'], bonus_cols['early_reason'], 200)
    ]

    for base_col, override_col, default_val in bonus_overrides:
        parts.append(
            _build_bonus_with_override_formula(
                base_col,
                override_col,
                row,
                default_val))

    return '=' + '+'.join(parts)


def process_kollel_attendance(
        input_file: str,
        output_file: str,
        working_days: int) -> bool:
    """
    Process Kollel attendance data and generate scholarship calculations.

    This function reads attendance data from an Excel file, processes it to calculate
    scholarships, and generates a new Excel file with detailed calculations and
    additional manual input columns.

    Args:
        input_file (str): Path to input Excel file containing attendance data
        output_file (str): Path where the output Excel file will be saved
        working_days (int): Number of working days in the period

    Returns:
        bool: True if processing was successful, False otherwise

    Raises:
        ValueError: If required columns are missing in the input file
        Exception: For other processing errors
    """
    try:
        df = pd.read_excel(input_file, dtype={
            'זהות': str,
            'שם משפחה': str,
            'שם פרטי': str,
            'רצופות': str
        })
        df = df.dropna(how='all')

        for col in ['כניסה', 'יציאה']:
            df[col] = pd.to_datetime(df[col])

        first_date = df['כניסה'].iloc[0]
        month_number = first_date.month
        year = first_date.year

        month_name = get_hebrew_month_name(month_number)
        output_file = f"מלגות חודשיות {month_name} {year}.xlsx"

        df['סך שעות'] = df['יציאה'] - df['כניסה']
        df['תאריך'] = df['כניסה'].dt.date
        df['שעת כניסה'] = df['כניסה'].dt.time
        df['שעת יציאה'] = df['יציאה'].dt.time
        df['סדר'] = df['שעת כניסה'].apply(
            lambda x: 'בוקר' if x.hour < 12 else 'צהריים')

        calculator = KollelScholarship()
        results = [
            calculator.calculate_student_scholarship(student_data, working_days)
            for _, student_data in df.groupby('זהות')
        ]

        results_df = pd.DataFrame(results)

        # Add "reason" columns for bonuses (to allow manual override)
        results_df['אנס דרגה 1'] = ''
        results_df['אנס דרגה 2'] = ''
        results_df['אנס נוכחות מושלמת'] = ''
        results_df['אנס הגעה מוקדמת'] = ''

        # Remove detailed statistics columns (morning_, afternoon_) -
        # they're only confusing in the summary
        stats_cols_to_remove = [col for col in results_df.columns if col.startswith(
            'בוקר_') or col.startswith('צהריים_')]
        results_df = results_df.drop(columns=stats_cols_to_remove)

        # Reorder columns - warnings right after name
        base_cols = ['מספר זהות', 'שם מלא', 'התראות']
        other_cols = [
            col for col in results_df.columns if col not in base_cols]
        results_df = results_df[base_cols + other_cols]

        # Add manual grade entry columns
        manual_columns = [
            'ציון חבורה', 'חבורה',
            'ציון מוסר', 'מוסר',
            'ציון סיכומים', 'סיכומים',
            'ציון מבחן הלכה', 'מבחן הלכה',
            'ציון מבחן שס', 'מבחן שס'
        ]
        for col in manual_columns:
            results_df[col] = ''

        results_df['סך סופי'] = 0

        # Replace all NaN values with empty string or 0
        results_df = results_df.fillna('')

        results_df.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        # Set right-to-left direction (RTL)
        ws.sheet_view.rightToLeft = True

        headers = [cell.value for cell in ws[1]]

        try:
            total_col = get_column_letter(headers.index('סך הכל') + 1)

            chabura_grade_col = get_column_letter(
                headers.index('ציון חבורה') + 1)
            musar_grade_col = get_column_letter(headers.index('ציון מוסר') + 1)
            sikumim_grade_col = get_column_letter(
                headers.index('ציון סיכומים') + 1)
            halacha_grade_col = get_column_letter(
                headers.index('ציון מבחן הלכה') + 1)
            shas_grade_col = get_column_letter(
                headers.index('ציון מבחן שס') + 1)

            chabura_col = get_column_letter(headers.index('חבורה') + 1)
            musar_col = get_column_letter(headers.index('מוסר') + 1)
            sikumim_col = get_column_letter(headers.index('סיכומים') + 1)
            halacha_col = get_column_letter(headers.index('מבחן הלכה') + 1)
            shas_col = get_column_letter(headers.index('מבחן שס') + 1)

            tier1_col = get_column_letter(headers.index('תוספת דרגה 1') + 1)
            tier1_reason_col = get_column_letter(
                headers.index('אנס דרגה 1') + 1)

            tier2_col = get_column_letter(headers.index('תוספת דרגה 2') + 1)
            tier2_reason_col = get_column_letter(
                headers.index('אנס דרגה 2') + 1)

            perfect_col = get_column_letter(headers.index('נוכחות מושלמת') + 1)
            perfect_reason_col = get_column_letter(
                headers.index('אנס נוכחות מושלמת') + 1)

            early_col = get_column_letter(headers.index('הגעה מוקדמת') + 1)
            early_reason_col = get_column_letter(
                headers.index('אנס הגעה מוקדמת') + 1)

            final_col = get_column_letter(headers.index('סך סופי') + 1)
        except ValueError as e:
            raise ValueError(f"לא נמצאה אחת העמודות הנדרשות: {str(e)}")

        for row in range(2, ws.max_row + 1):
            # Grade formulas
            ws[f'{chabura_col}{row}'] = '=' + \
                _build_grade_formula(chabura_grade_col, row, 150)
            ws[f'{musar_col}{row}'] = '=' + \
                _build_grade_formula(musar_grade_col, row, 100)
            ws[f'{sikumim_col}{row}'] = '=' + \
                _build_grade_formula(sikumim_grade_col, row, 150)
            ws[f'{halacha_col}{row}'] = '=' + \
                _build_grade_formula(halacha_grade_col, row, 200)
            ws[f'{shas_col}{row}'] = '=' + \
                _build_grade_formula(shas_grade_col, row, 390)

            # Final total formula
            grade_cols = {
                'chabura': chabura_col,
                'musar': musar_col,
                'sikumim': sikumim_col,
                'halacha': halacha_col,
                'shas': shas_col
            }

            bonus_cols = {
                'tier1': tier1_col,
                'tier1_reason': tier1_reason_col,
                'tier2': tier2_col,
                'tier2_reason': tier2_reason_col,
                'perfect': perfect_col,
                'perfect_reason': perfect_reason_col,
                'early': early_col,
                'early_reason': early_reason_col
            }

            ws[f'{final_col}{row}'] = _build_final_total_formula(
                total_col, grade_cols, bonus_cols, row)

        wb.save(output_file)

        # Add detailed sheets for each student
        try:
            add_detailed_sheets(output_file, df, results, working_days)
        except Exception as e:
            print(f"Warning: Could not add detailed sheets: {e}")
            # Continue even if adding sheets failed

        # Format the summary sheet
        try:
            format_summary_sheet(output_file)
        except Exception as e:
            print(f"Warning: Could not format summary sheet: {e}")
            # Continue even if formatting failed

        return True

    except Exception as e:
        print(f"An error occurred: {e}")
        messagebox.showerror("שגיאה", f"אירעה שגיאה: {str(e)}")
        return False
